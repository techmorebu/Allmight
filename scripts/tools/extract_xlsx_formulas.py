#!/usr/bin/env python3
"""
Extract all formulas from an .xlsx/.xlsm workbook.

Outputs:
- CSV: sheet, cell, formula
- JSON: {sheet: [{cell, formula}, ...], ...}

Loads workbook with data_only=False to preserve formulas.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook  # type: ignore


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Path to .xlsx/.xlsm")
    p.add_argument("--outdir", default="outputs/formulas", help="Output directory")
    args = p.parse_args()

    in_path = Path(args.input)
    outdir = Path(args.outdir)

    if not in_path.exists():
        raise FileNotFoundError(f"Missing workbook: {in_path}")

    outdir.mkdir(parents=True, exist_ok=True)
    csv_path = outdir / f"{in_path.stem}_formulas.csv"
    json_path = outdir / f"{in_path.stem}_formulas.json"

    wb = load_workbook(in_path, data_only=False, read_only=False)

    all_rows: List[List[str]] = []
    by_sheet: Dict[str, List[Dict[str, str]]] = {}

    for ws in wb.worksheets:
        by_sheet[ws.title] = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    all_rows.append([ws.title, cell.coordinate, v])
                    by_sheet[ws.title].append({"cell": cell.coordinate, "formula": v})

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "cell", "formula"])
        w.writerows(all_rows)

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(by_sheet, f, indent=2)

    print(f"Workbook: {in_path}")
    print(f"Sheets: {len(wb.worksheets)}")
    print(f"Formulas found: {len(all_rows)}")
    print(f"Wrote CSV: {csv_path}")
    print(f"Wrote JSON: {json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
